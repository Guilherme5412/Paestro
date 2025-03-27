from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import io
from datetime import datetime
import re

def export_to_excel(classes, attendance_status, observations, html_content=None, current_user=None, periodo=None, escola_nome=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "LISTA DE PRESENÇA"  # Título em maiúsculas

    # Cabeçalho com informações gerais
    header_rows = [
        ("UNIDADE:", (escola_nome or "NÃO INFORMADO").upper()),
        ("RESPONSÁVEIS:", (current_user or "NÃO INFORMADO").upper()),
        ("PERÍODO:", (periodo or "NÃO INFORMADO").upper()),
        ("DATA E HORA:", datetime.now().strftime('%d/%m/%Y %H:%M'))
    ]

    for i, (label, value) in enumerate(header_rows, start=1):
        ws[f'A{i}'] = label
        ws[f'A{i}'].font = Font(bold=True)
        ws[f'B{i}'] = value
        ws.merge_cells(f'B{i}:D{i}')

    current_row = len(header_rows) + 2

    if not classes:
        ws.merge_cells(f"A{current_row}:D{current_row}")
        ws[f"A{current_row}"] = "NENHUMA CHAMADA SALVA ENCONTRADA"
        ws[f"A{current_row}"].font = Font(italic=True)
    else:
        for turma, alunos in classes.items():
            # Cabeçalho da turma
            ws.merge_cells(f"A{current_row}:D{current_row}")
            ws[f"A{current_row}"] = f"TURMA: {turma.upper()}"
            ws[f"A{current_row}"].font = Font(bold=True, size=12)
            current_row += 1

            # Cabeçalho da tabela
            headers = ["ALUNO", "PRESENÇA", "OBSERVAÇÃO"]
            for col, header in enumerate(headers, start=1):
                ws.cell(row=current_row, column=col).value = header
                ws.cell(row=current_row, column=col).font = Font(bold=True)
            current_row += 1

            # Dados dos alunos
            for aluno in alunos:
                ws.cell(row=current_row, column=1).value = aluno
                ws.cell(row=current_row, column=2).value = attendance_status.get(turma, {}).get(aluno, "P")
                ws.cell(row=current_row, column=3).value = observations.get(turma, {}).get(aluno, "")
                current_row += 1

            current_row += 1  # Espaço entre turmas

    # Ajuste de colunas
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 40

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def get_excel_filename(escola_nome=None, periodo=None, current_user=None):
    data = datetime.now().strftime('%d-%m-%Y')
    
    # Formata o nome da escola (remove caracteres especiais e converte para maiúsculas)
    if escola_nome:
        # Remove caracteres não alfanuméricos exceto espaços e underscores
        nome_escola = re.sub(r'[^\w\s-]', '', escola_nome).strip()
        # Substitui espaços por underscores e converte para maiúsculas
        nome_escola = nome_escola.replace(' ', '_').upper()
    else:
        nome_escola = "ESCOLA_INDEFINIDA"
    
    # Formata o nome da dupla
    if current_user:
        # Remove caracteres não alfanuméricos exceto espaços e underscores
        nome_dupla = re.sub(r'[^\w\s-]', '', current_user).strip()
        # Substitui espaços por underscores e converte para maiúsculas
        nome_dupla = nome_dupla.replace(' ', '_').upper()
    else:
        nome_dupla = "DUPLA_INDEFINIDA"
    
    # Formata o período
    if periodo:
        periodo_formatado = periodo.upper()
    else:
        periodo_formatado = "PERIODO_INDEFINIDO"
    
    return f"{nome_escola}_{data}_{nome_dupla}_{periodo_formatado}.XLSX"